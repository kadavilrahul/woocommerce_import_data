import requests
import time
import openpyxl
import os
from urllib.parse import urljoin, urlparse

try:
    from config import CONSUMER_KEY, CONSUMER_SECRET, SITE_URL
except ImportError:
    CONSUMER_KEY = input("Enter your WooCommerce Consumer Key: ")
    CONSUMER_SECRET = input("Enter your WooCommerce Consumer Secret: ")
    SITE_URL = input("Enter your WooCommerce Site URL (e.g., https://example.com): ").strip()

# Enhanced URL validation and formatting
def format_site_url(url):
    """Format and validate the site URL."""
    url = url.strip('/')
    parsed = urlparse(url)
    
    if not parsed.scheme:
        url = 'https://' + url
    
    print(f"Formatted URL: {url}")  # Debug output
    return url

SITE_URL = format_site_url(SITE_URL)
EXCEL_FILE = "products.xlsx"
PAGE_PROPERTY_FILE = "current_page.txt"

def get_current_page():
    """Retrieve the current page from the property file."""
    try:
        with open(PAGE_PROPERTY_FILE, "r") as file:
            return int(file.read().strip())
    except FileNotFoundError:
        return 1

def save_current_page(page):
    """Save the current page number to the property file."""
    with open(PAGE_PROPERTY_FILE, "w") as file:
        file.write(str(page))

def fetch_titles(page):
    """Fetch product titles from WooCommerce API."""
    try:
        # Construct API URL
        api_url = f"{SITE_URL}/wp-json/wc/v3/products"
        
        params = {
            'page': page,
            'per_page': 50,
            'consumer_key': CONSUMER_KEY,
            'consumer_secret': CONSUMER_SECRET
        }
        
        # Debug information
        print(f"\nDebug Information:")
        print(f"Base URL: {SITE_URL}")
        print(f"API URL: {api_url}")
        print(f"Full URL with params: {api_url}?page={page}&per_page=50&consumer_key=<hidden>&consumer_secret=<hidden>")
        
        response = requests.get(api_url, params=params)
        
        # More debug information
        print(f"Response Status Code: {response.status_code}")
        if response.status_code != 200:
            print(f"Response Text: {response.text[:200]}...")  # Print first 200 chars of response
        
        response.raise_for_status()
        products = response.json()
        return [product["name"] for product in products], len(products) == 50
        
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        print(f"Full error details: {str(e)}")
        return [], False

def write_titles_to_excel(titles):
    """Write titles to an Excel file."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    
    sheet = wb.active
    last_row = sheet.max_row
    
    if last_row == 1 and not sheet['A1'].value:
        last_row = 0
        
    for idx, title in enumerate(titles, start=last_row + 1):
        sheet.cell(row=idx, column=1, value=title)
    
    wb.save(EXCEL_FILE)

def fetch_woocommerce_product_titles(num_products=None):
    """Main function to fetch product titles and save them to Excel."""
    current_page = get_current_page()
    total_products = 0
    
    try:
        while True:
            titles, has_more = fetch_titles(current_page)
            
            if not titles:
                print("No more products found or error occurred.")
                break
                
            write_titles_to_excel(titles)
            print(f"Page {current_page}: Imported {len(titles)} product titles.")
            save_current_page(current_page)
            
            total_products += len(titles)
            if num_products is not None and total_products >= num_products:
                print(f"Successfully imported {total_products} products.")
                break
                
            if not has_more:
                print("All products have been fetched and listed in the Excel file.")
                break
                
            current_page += 1
            print("Waiting 15 seconds before next request...")
            time.sleep(15)
            
    except Exception as e:
        print(f"An error occurred: {e}")
        return False
    
    return True

def main():
    print("Welcome to the Product Title Importer!")
    print("\nIMPORTANT: Please enter the site URL in one of these formats:")
    print("- example.com")
    print("- https://example.com")
    print("- http://example.com")
    
    while True:
        choice = input("\nDo you want to import a specific number of products? (yes/no): ").strip().lower()
        
        if choice == 'yes':
            try:
                num_products = int(input("Enter the number of products to import: "))
                if num_products > 0:
                    print(f"Importing {num_products} products...")
                    if fetch_woocommerce_product_titles(num_products):
                        break
                else:
                    print("Please enter a positive number.")
            except ValueError:
                print("Invalid input. Please enter a valid number.")
            except KeyboardInterrupt:
                print("\nOperation cancelled by user.")
                break
        elif choice == 'no':
            print("Importing all products...")
            if fetch_woocommerce_product_titles():
                break
        else:
            print("Please enter 'yes' or 'no'.")

if __name__ == "__main__":
    main()